"""
转介绍规划表生成器
读取BI源数据 → 提取月度小计 → 写入标准规划表
"""
import re
import sys
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any

# 让父级 src 目录可导入
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from data_processor import XlsxReader, DataProcessor
from agents.planning_config import (
    CALIBERS, METRICS, CALIBER_HEADERS, HISTORY_MONTHS,
    MONTHLY_TARGETS, calculate_time_progress, get_current_month_key,
    get_targets_for_month, OUTPUT_PATH,
)

logger = logging.getLogger(__name__)

# openpyxl 用于写入（比 xlsxwriter 更适合"更新已有文件"的场景）
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("需要 openpyxl: pip install openpyxl")


class PlanningGenerator:
    """转介绍规划表生成器"""

    def __init__(self, source_path: str, report_date: datetime = None):
        self.source_path = Path(source_path)
        self.report_date = report_date or datetime.now()
        self.month_key = get_current_month_key(self.report_date)
        self.targets = get_targets_for_month(self.month_key)
        self.time_progress = calculate_time_progress(self.report_date)

        # 提取源数据
        self.monthly_data: Dict[str, Dict] = {}
        self._extract_data()

    def _extract_data(self):
        """从BI源文件提取月度小计数据"""
        reader = XlsxReader(str(self.source_path))
        processor = DataProcessor(reader)
        result = processor.process()
        self.monthly_data = result["monthly_summaries"]
        logger.info(f"提取到 {len(self.monthly_data)} 个月份的数据: {sorted(self.monthly_data.keys())}")

    def generate(self, output_path: str = None) -> Path:
        """生成规划表 xlsx"""
        out = Path(output_path) if output_path else OUTPUT_PATH
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active

        # Sheet 名 = 当前月 + "转介绍参与率"
        month_num = int(self.month_key[-2:])
        ws.title = f"{month_num}月转介绍参与率"

        # --- 构建表格 ---
        self._write_header(ws)
        data_start_row = 3
        row = data_start_row
        row = self._write_monthly_data(ws, row)
        row = self._write_growth_row(ws, row)
        row += 1  # 空行
        row = self._write_progress_section(ws, row)
        row += 1
        row = self._write_problem_section(ws, row)
        row += 1
        row = self._write_target_section(ws, row)
        row += 1
        row = self._write_participation_section(ws, row)
        row += 1
        row = self._write_action_section(ws, row)
        row += 2
        # 底部重复 202509 数据行（与原表结构一致）
        row = self._write_bottom_reference(ws, row)

        # 调整列宽
        self._auto_column_width(ws)
        # 应用样式
        self._apply_styles(ws, data_start_row)

        wb.save(str(out))
        logger.info(f"规划表已生成: {out}")
        return out

    # ========== 写入各部分 ==========

    def _write_header(self, ws):
        """写入表头 (2行: 口径分组 + 指标名)"""
        # 第1行: 口径分组
        row1 = ["REFER"]
        for cal in CALIBERS:
            header = CALIBER_HEADERS[cal]
            row1.extend([header] * len(METRICS))
        for c, val in enumerate(row1, 1):
            ws.cell(row=1, column=c, value=val)

        # 第2行: 指标名
        row2 = ["REFER"]
        metric_labels_total = [
            "注册Register", "预约Appt.", "出席Show up", "付费 Paid",
            "美金金额 USD", "注册付费率 leads to pay %", "预约率 Appt%",
            "预约出席率show up%", "出席付费率 show up to pay%"
        ]
        metric_labels_sub = ["注册", "预约", "出席", "付费", "美金金额",
                             "注册付费率", "预约率", "预约出席率", "出席付费率"]
        row2.extend(metric_labels_total)
        for _ in range(3):
            row2.extend(metric_labels_sub)
        for c, val in enumerate(row2, 1):
            ws.cell(row=2, column=c, value=val)

    def _write_monthly_data(self, ws, start_row: int) -> int:
        """写入各月份数据行"""
        row = start_row
        for month in HISTORY_MONTHS:
            data = self.monthly_data.get(month)
            if data is None:
                continue
            ws.cell(row=row, column=1, value=month)
            col = 2
            for cal in CALIBERS:
                prefix = f"{cal}_"
                for metric in METRICS:
                    key = prefix + metric
                    ws.cell(row=row, column=col, value=data.get(key))
                    col += 1
            row += 1
        return row

    def _write_growth_row(self, ws, row: int) -> int:
        """写入环比GAP行"""
        ws.cell(row=row, column=1, value="环比GAP")
        months = [m for m in HISTORY_MONTHS if m in self.monthly_data]
        if len(months) >= 2:
            curr = self.monthly_data[months[-1]]
            prev = self.monthly_data[months[-2]]
            col = 2
            for cal in CALIBERS:
                prefix = f"{cal}_"
                for metric in METRICS:
                    key = prefix + metric
                    cv, pv = curr.get(key), prev.get(key)
                    if cv is not None and pv is not None and pv != 0:
                        ws.cell(row=row, column=col, value=(cv - pv) / pv)
                    col += 1
        return row + 1

    def _write_progress_section(self, ws, row: int) -> int:
        """写入时间进度与目标完成度"""
        targets = self.targets
        if not targets:
            return row

        curr_month = self.month_key
        data = self.monthly_data.get(curr_month, {})

        bm = self.time_progress
        total_target = targets.get("总标", 0)
        unit_price = targets.get("客单价", 850)
        conv_rate = targets.get("转率目标", 0.23)

        # 反推目标
        paid_target = total_target / unit_price if unit_price else 0
        reg_target = paid_target / conv_rate if conv_rate else 0
        # 预约目标 = 注册目标 × 预约率(取上月数据)
        months = [m for m in HISTORY_MONTHS if m in self.monthly_data]
        prev_data = self.monthly_data.get(months[-2], {}) if len(months) >= 2 else {}
        appt_rate = prev_data.get("总计_预约率", 0.77) or 0.77
        show_rate = prev_data.get("总计_预约出席率", 0.66) or 0.66
        appt_target = reg_target * appt_rate
        show_target = appt_target * show_rate

        # 当前完成
        reg_done = data.get("总计_注册", 0) or 0
        appt_done = data.get("总计_预约", 0) or 0
        show_done = data.get("总计_出席", 0) or 0
        paid_done = data.get("总计_付费", 0) or 0
        amt_done = data.get("总计_美金金额", 0) or 0
        conv_done = data.get("总计_注册付费率", 0) or 0

        # Today BM
        ws.cell(row=row, column=1, value="Today BM")
        for c in range(2, 8):
            ws.cell(row=row, column=c, value=bm)
        row += 1

        # 目标行
        month_label = f"{int(curr_month[-2:])}月目标 Target"
        ws.cell(row=row, column=1, value=month_label)
        targets_row = [reg_target, appt_target, show_target, paid_target, total_target, conv_rate]
        for i, v in enumerate(targets_row):
            ws.cell(row=row, column=2 + i, value=v)
        row += 1

        # 完成行
        month_label2 = f"{int(curr_month[-2:])}月完成 Achievement"
        ws.cell(row=row, column=1, value=month_label2)
        done_row = [reg_done, appt_done, show_done, paid_done, amt_done, conv_done]
        for i, v in enumerate(done_row):
            ws.cell(row=row, column=2 + i, value=v)
        row += 1

        # 效率进度
        month_label3 = f"{int(curr_month[-2:])}月效率进度 BM"
        ws.cell(row=row, column=1, value=month_label3)
        eff_row = []
        for done, target in zip(done_row[:5], targets_row[:5]):
            if target and target != 0:
                eff_row.append(done / target)
            else:
                eff_row.append(None)
        # 转化率的效率进度
        if conv_rate and conv_rate != 0:
            eff_row.append(conv_done / conv_rate)
        else:
            eff_row.append(None)
        for i, v in enumerate(eff_row):
            ws.cell(row=row, column=2 + i, value=v)
        row += 1

        # 目标GAP
        ws.cell(row=row, column=1, value="目标 GAP")
        for i, v in enumerate(eff_row):
            gap = (v - bm) if v is not None else None
            ws.cell(row=row, column=2 + i, value=gap)
        row += 1

        return row

    def _write_problem_section(self, ws, row: int) -> int:
        """写入当前问题分析"""
        targets = self.targets
        if not targets:
            return row

        curr_month = self.month_key
        data = self.monthly_data.get(curr_month, {})

        total_target = targets.get("总标", 0)
        unit_price = targets.get("客单价", 850)
        conv_rate = targets.get("转率目标", 0.23)

        paid_target = total_target / unit_price if unit_price else 0
        reg_target = paid_target / conv_rate if conv_rate else 0

        amt_done = data.get("总计_美金金额", 0) or 0
        paid_done = data.get("总计_付费", 0) or 0
        show_done = data.get("总计_出席", 0) or 0
        appt_done = data.get("总计_预约", 0) or 0

        # 预约/出席目标
        months = [m for m in HISTORY_MONTHS if m in self.monthly_data]
        prev_data = self.monthly_data.get(months[-2], {}) if len(months) >= 2 else {}
        appt_rate = prev_data.get("总计_预约率", 0.77) or 0.77
        show_rate = prev_data.get("总计_预约出席率", 0.66) or 0.66
        appt_target = reg_target * appt_rate
        show_target = appt_target * show_rate

        actual_unit_price = amt_done / paid_done if paid_done else 0

        problems = [
            ("业绩缺口", amt_done - total_target),
            ("客单价", actual_unit_price),
            ("单量缺口", paid_done - paid_target),
            ("出席缺口", show_done - show_target),
            ("预约缺口", appt_done - appt_target),
        ]

        for label, value in problems:
            ws.cell(row=row, column=1, value="当前问题")
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=value)
            row += 1

        return row

    def _write_target_section(self, ws, row: int) -> int:
        """写入目标配置区域"""
        targets = self.targets
        if not targets:
            return row

        total_target = targets.get("总标", 0)
        unit_price = targets.get("客单价", 850)
        conv_rate = targets.get("转率目标", 0.23)
        paid_target = total_target / unit_price if unit_price else 0
        reg_target = paid_target / conv_rate if conv_rate else 0

        month_num = int(self.month_key[-2:])

        ws.cell(row=row, column=2, value="目标")
        row += 1

        items = [
            (f"{month_num}月总标", total_target),
            (f"{month_num}月客单价", unit_price),
            ("单量目标", paid_target),
            ("转率目标", conv_rate),
            ("例子目标", reg_target),
        ]
        for label, value in items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        return row

    def _write_participation_section(self, ws, row: int) -> int:
        """写入转介绍参与率区域（结构框架，具体数据留空手填）"""
        month_num = int(self.month_key[-2:])
        prev_month = month_num - 1 if month_num > 1 else 12

        ws.cell(row=row, column=8, value=f"{prev_month}月转介绍参与率")
        ws.cell(row=row, column=9, value=f"{month_num}月转介绍参与率")
        row += 1

        periods = ["0-30天", "30-60天", "60天-90天", "90天以上"]
        for p in periods:
            ws.cell(row=row, column=7, value=p)
            row += 1

        return row

    def _write_action_section(self, ws, row: int) -> int:
        """写入行动计划与口径评价"""
        action_text = (
            '\uff081\uff09\u6708\u672b\u4e09\u5929\u51fa\u5e2d\u8bfe\u91cf\u6fc0\u52b1\uff0c\u76ee\u6807\u5012\u63a8\u9488\u5bf9\u5185\u573a\u505a\u201c\u51fa\u5e2d\u6fc0\u52b1\u201d\uff1b\n'
            '\uff082\uff09\u6708\u672b\u4e09\u5929\u5206\u914dTMK\u516c\u6d77\u6570\u636e\uff0c\u6309\u73b0\u6709\u516c\u6d77\u6570\u636e\u9080\u7ea6\u6548\u73873.5%\uff0c\u9700\u62e8\u6253\u5b8c\u62102000+\u5386\u53f2\u7528\u6237\uff1b\n'
            '\uff083\uff09\u6708\u672b\u4e09\u5929\u5355\u91cf\u6fc0\u52b1\uff0c\u505a\u4e2a\u4eba\u8f6c\u4ecb\u7ecd\u5355\u91cf\u9636\u68af\u6fc0\u52b1\uff1b'
        )
        ws.cell(row=row, column=1, value=action_text)

        # 口径评价标签
        labels = ["CC", "SS", "LP", "宽口径"]
        for i, label in enumerate(labels):
            ws.cell(row=row, column=8 + i, value=label)
        row += 1

        evaluations = ["没问题", "没问题，预约有点儿小问题", "问题不大", "开源差且有注水"]
        for i, ev in enumerate(evaluations):
            ws.cell(row=row, column=8 + i, value=ev)
        row += 1

        return row

    def _write_bottom_reference(self, ws, row: int) -> int:
        """写入底部参考数据 (第一个月的数据，与原表结构一致)"""
        # 口径表头
        cal_row1 = []
        for cal in CALIBERS:
            label = "总计" if cal == "总计" else ("CC窄口径" if cal == "CC窄口径" else ("SS窄口径" if cal == "SS窄口径" else "其它"))
            cal_row1.extend([label] * len(METRICS))
        for c, val in enumerate(cal_row1, 1):
            ws.cell(row=row, column=c, value=val)
        row += 1

        # 指标名
        metric_labels = ["注册", "预约", "出席", "付费", "美金金额",
                         "注册付费率", "预约率", "预约出席率", "出席付费率"]
        metric_row = []
        for _ in CALIBERS:
            metric_row.extend(metric_labels)
        for c, val in enumerate(metric_row, 1):
            ws.cell(row=row, column=c, value=val)
        row += 1

        # 第一个月数据
        first_month = HISTORY_MONTHS[0] if HISTORY_MONTHS else None
        if first_month and first_month in self.monthly_data:
            data = self.monthly_data[first_month]
            col = 1
            for cal in CALIBERS:
                prefix = f"{cal}_"
                for metric in METRICS:
                    ws.cell(row=row, column=col, value=data.get(prefix + metric))
                    col += 1
        row += 1

        return row

    # ========== 样式 ==========

    def _apply_styles(self, ws, data_start_row: int):
        """应用基础格式"""
        # 表头样式
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        for row in ws.iter_rows(min_row=1, max_row=2, max_col=ws.max_column):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 百分比列格式
        pct_metrics = {"注册付费率", "预约率", "预约出席率", "出席付费率"}
        for row in ws.iter_rows(min_row=data_start_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                # 根据表头判断是否为百分比
                header = ws.cell(row=2, column=cell.column).value or ""
                if any(p in header for p in ["率", "%", "BM", "GAP", "进度"]):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'
                elif isinstance(cell.value, float) and abs(cell.value) > 100:
                    cell.number_format = '#,##0'
                elif isinstance(cell.value, int):
                    cell.number_format = '#,##0'

        # 特殊行样式
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column):
            first_val = row[0].value
            if first_val and isinstance(first_val, str):
                if "环比" in first_val:
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.0%'
                elif "BM" in first_val or "效率" in first_val or "GAP" in first_val:
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.0%'

    def _auto_column_width(self, ws):
        """自动调整列宽"""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    cell_len = len(str(cell.value))
                    max_length = max(max_length, min(cell_len, 18))
            ws.column_dimensions[col_letter].width = max(max_length + 2, 8)


def generate_planning_report(source_path: str, output_path: str = None,
                              report_date: datetime = None) -> Path:
    """便捷入口函数"""
    gen = PlanningGenerator(source_path, report_date=report_date)
    return gen.generate(output_path)
